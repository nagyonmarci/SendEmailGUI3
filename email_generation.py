import os
import sys
import subprocess
import tempfile
import openpyxl

IS_WINDOWS = sys.platform == 'win32'
IS_MAC = sys.platform == 'darwin'

if IS_WINDOWS:
    import pythoncom
    import win32com.client as win32

DEFAULT_HTML_BODY_FIRST = """<font style="font-family:Verdana" size="10pt" color="#184879">"""
DEFAULT_HTML_BODY_LAST = """</font>"""

stop_generation = False
generation_in_progress = False

_OUTLOOK_MAC_SCRIPT = """on run argv
    set htmlPath to item 1 of argv
    set msgSubject to item 2 of argv
    set toAddr to item 3 of argv
    set toName to item 4 of argv
    set ccAddr to item 5 of argv

    set htmlContent to (do shell script "cat " & quoted form of htmlPath)

    tell application "Microsoft Outlook"
        set newMsg to make new outgoing message
        set subject of newMsg to msgSubject
        make new |html body| at newMsg with properties {content:htmlContent}
        make new to recipient at newMsg with properties {email address:{address:toAddr, name:toName}}
        if ccAddr is not "" then
            make new cc recipient at newMsg with properties {email address:{address:ccAddr}}
        end if
        set i to 6
        repeat while i <= (count of argv)
            set attPath to item i of argv
            if attPath is not "" then
                make new attachment at newMsg with properties {file:(POSIX file attPath)}
            end if
            set i to i + 1
        end repeat
        open newMsg
        activate
    end tell
end run"""


def _create_email_mac(recipient_name, recipient_email, cc_email, subject, html_content, attachments):
    full_html = DEFAULT_HTML_BODY_FIRST + html_content + DEFAULT_HTML_BODY_LAST

    html_tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8')
    html_tmp.write(full_html)
    html_tmp.close()

    script_tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.applescript', delete=False, encoding='utf-8')
    script_tmp.write(_OUTLOOK_MAC_SCRIPT)
    script_tmp.close()

    cmd = [
        'osascript', script_tmp.name,
        html_tmp.name,
        subject,
        recipient_email,
        recipient_name or '',
        cc_email or '',
    ] + attachments

    result = subprocess.run(cmd, capture_output=True)
    if result.returncode != 0:
        print(f"AppleScript hiba: {result.stderr.decode()}")

    os.unlink(script_tmp.name)


def close_all_open_email_windows():
    if IS_WINDOWS:
        outlook = win32.Dispatch("Outlook.Application")
        inspectors = outlook.Inspectors
        for inspector in list(inspectors):
            item = inspector.CurrentItem
            try:
                item.Close(0)
            except Exception as e:
                print(f"Hiba történt az ablak bezárása közben: {e}")
    else:
        subprocess.run(
            ['osascript', '-e', 'tell application "Microsoft Outlook" to close every window'],
            capture_output=True
        )


def stop_email_generation():
    global stop_generation
    stop_generation = True


def generate_emails(table_filename, attachment_dir, sheet_name_entry, subject_entry, html_body_text):
    if IS_WINDOWS:
        pythoncom.CoInitialize()
    try:
        global stop_generation, generation_in_progress
        generation_in_progress = True
        workbook = openpyxl.load_workbook(table_filename.get())
        sheet = workbook[sheet_name_entry.get()]

        if IS_WINDOWS:
            outlook = win32.Dispatch('outlook.application')

        for i in range(2, sheet.max_row + 1):
            if stop_generation:
                print("E-mail generálás megállítva.")
                break

            attachments_raw = sheet.cell(row=i, column=1).value.split(';')
            recipient_name = sheet.cell(row=i, column=2).value
            recipient_email = sheet.cell(row=i, column=3).value
            cc_email = sheet.cell(row=i, column=4).value
            subject = subject_entry.get()
            if hasattr(html_body_text, 'get_html'):
                html_content = html_body_text.get_html()
            else:
                html_content = html_body_text.get("1.0", "end-1c").replace("\n", "<br>")

            if IS_WINDOWS:
                mail = outlook.CreateItem(0)
                mail.Display(False)
                if recipient_name:
                    mail.To = f"{recipient_name} <{recipient_email}>"
                else:
                    mail.To = recipient_email
                if cc_email:
                    mail.CC = cc_email
                mail.Subject = subject
                current_body = mail.HTMLBody
                mail.HTMLBody = DEFAULT_HTML_BODY_FIRST + html_content + DEFAULT_HTML_BODY_LAST + current_body
                for attachment in attachments_raw:
                    if attachment:
                        attachment_path = os.path.join(attachment_dir.get(), attachment.strip())
                        if os.path.exists(attachment_path):
                            mail.Attachments.Add(attachment_path)
                mail.Display()
            else:
                full_attachments = []
                for attachment in attachments_raw:
                    if attachment:
                        attachment_path = os.path.join(attachment_dir.get(), attachment.strip())
                        if os.path.exists(attachment_path):
                            full_attachments.append(attachment_path)
                _create_email_mac(recipient_name, recipient_email, cc_email,
                                  subject, html_content, full_attachments)

    except Exception as e:
        print(f"Hiba történt: {e}")
    finally:
        if IS_WINDOWS:
            pythoncom.CoUninitialize()
        if 'workbook' in locals():
            workbook.close()
        generation_in_progress = False
        stop_generation = False
